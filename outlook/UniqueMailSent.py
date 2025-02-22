import win32com.client
import openpyxl
import datetime

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
sent_items = outlook.GetDefaultFolder(5)  # Получаем папку "Отправленные"

# Получаем все письма из папки "Отправленные" за 2023 год
messages = sent_items.Items.Restrict("[SentOn] >= '01/01/2023' AND [SentOn] <= '12/31/2023'")

unique_recipients = set()
checked_emails = 0

# Собираем уникальные адреса получателей и их имена (если доступно)
for message in messages:
    recipients = message.Recipients
    for recipient in recipients:
        email_addr = recipient.Address
        name = recipient.Name if hasattr(recipient, 'Name') else None
        if email_addr not in unique_recipients:
            unique_recipients.add((email_addr, name))
    checked_emails += 1
# Создаем новый документ Excel
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Recipients"

# Записываем уникальные адреса получателей и их имена в документ Excel
row_num = 1
for email_addr, name in unique_recipients:
    if name:
        sheet.cell(row=row_num, column=1, value=name)
    else:
        sheet.cell(row=row_num, column=1, value=email_addr)
    sheet.cell(row=row_num, column=2, value=email_addr)
    row_num += 1

# Устанавливаем заголовки для столбцов
sheet.cell(row=1, column=1, value="Name")
sheet.cell(row=1, column=2, value="Email Address")

# Сохраняем документ Excel
workbook.save("unique_recipients_sent_2023.xlsx")
print(f"Проверено {checked_emails} писем.")